from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = Path(__file__).with_name("budget-variants.xlsx")


def style_header(row):
    for cell in row:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="24445C")
        cell.alignment = Alignment(vertical="center")


def set_widths(sheet, widths):
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


workbook = Workbook()
assumptions = workbook.active
assumptions.title = "Annahmen"
assumptions.append(["Eingabe", "Wert", "Hinweis"])
style_header(assumptions[1])

inputs = [
    ("Monatsbrutto Vollzeit Thomas", 4500, "Arbeitsannahme; durch WE AID und Thomas ersetzen"),
    ("Arbeitgeberaufschlag", 0.23, "Arbeitsannahme; echte Lohnnebenkosten einsetzen"),
    ("Stellenanteil", 0.60, "Planvorgabe"),
    ("Projektmonate", 10, "bpb 01.03.–31.12.2027; für Mercator auf 12 oder 18 ändern"),
    ("WE-AID-Anteil", 0.07, "Förderfähigkeit und Berechnungsbasis schriftlich klären"),
    ("Bestätigte Kofinanzierung", 0, "Nur schriftlich bestätigte liquide Mittel"),
]
for row in inputs:
    assumptions.append(row)

input_fill = PatternFill("solid", fgColor="FFF2CC")
for cell in assumptions["B"][1:]:
    cell.fill = input_fill
assumptions["B2"].number_format = '#,##0.00 "€"'
assumptions["B3"].number_format = "0.0%"
assumptions["B4"].number_format = "0%"
assumptions["B6"].number_format = "0.0%"
assumptions["B7"].number_format = '#,##0.00 "€"'
set_widths(assumptions, [36, 18, 65])
assumptions.freeze_panes = "A2"

budget = workbook.create_sheet("Budgetvarianten")
budget.append(["Kostenblock", "Schlank", "Kern", "Voll", "Gate"])
style_header(budget[1])

rows = [
    ("Thomas, 60%-Projektstelle inkl. AG-Kosten", "=Annahmen!B2*Annahmen!B4*Annahmen!B5*(1+Annahmen!B3)", "=B2", "=B2", "Neuer Projektvertrag; ALG I vorher schriftlich klären"),
    ("Zielgruppenpartner", 5000, 7500, 10000, "Rollenbrief und Angebote"),
    ("Evaluation", 4000, 8000, 12000, "Unabhängigkeit und Datenschutz"),
    ("Kontrollierte Reichweitentests", 2000, 5000, 8000, "Kein Microtargeting; Förderfähigkeit bestätigen"),
    ("Infrastruktur und Barrierefreiheit", 3000, 5000, 8000, "Nur projektbezogene Mehrkosten"),
    ("Veranstaltungen und Materialien", 2000, 4000, 6000, "Programmregeln und Angebote"),
    ("Reisekosten", 1000, 2000, 3000, "Förderrichtlinie anwenden"),
    ("Zwischensumme direkte Kosten", "=SUM(B2:B8)", "=SUM(C2:C8)", "=SUM(D2:D8)", ""),
    ("WE AID", "=B9*Annahmen!B6", "=C9*Annahmen!B6", "=D9*Annahmen!B6", "Basis und Förderfähigkeit klären"),
    ("Gesamtausgaben", "=SUM(B9:B10)", "=SUM(C9:C10)", "=SUM(D9:D10)", ""),
    ("Bestätigte Kofinanzierung", "=Annahmen!B7", "=Annahmen!B7", "=Annahmen!B7", "Keine Eigenleistung oder private Kredite"),
    ("Beantragter Förderbedarf", "=B11-B12", "=C11-C12", "=D11-D12", "Vor Einreichung auf Programmgrenze anpassen"),
]
for row in rows:
    budget.append(row)

for row in budget.iter_rows(min_row=2, max_row=13, min_col=2, max_col=4):
    for cell in row:
        cell.number_format = '#,##0.00 "€"'
for row_number in (9, 11, 13):
    for cell in budget[row_number]:
        cell.font = Font(bold=True)
budget.auto_filter.ref = "A1:E13"
budget.freeze_panes = "B2"
set_widths(budget, [44, 16, 16, 16, 62])

programs = workbook.create_sheet("Programmzuordnung")
programs.append(["Programm", "Budget verwenden?", "Regel"])
style_header(programs[1])
program_rows = [
    ("bpb-Modellförderung", "Ja, nach Richtlinienprüfung", "Offiziellen Ausgaben-/Finanzierungsplan verwenden; Werbe- und Verwaltungskosten bestätigen"),
    ("Stiftung Mercator", "Ja, als Alternative", "Keine identischen Kosten zugleich mit bpb finanzieren"),
    ("Fast Forward", "Nein", "Unrestricted grant; eigene USD-Finanzangaben im Formular"),
    ("Erasmus+ Jugendpartizipation", "Nein", "Ausschließlich aktuelle Pauschalen verwenden"),
    ("Bremer Kleinförderung", "Nein", "Eigenes Maximalbudget von 3.500 Euro"),
    ("Shuttleworth", "Nein", "Geschlossen; kein Intake 2026/2027"),
    ("NLnet", "Nein", "Nur technischer Open-Source-Scope; Thomas schreibt selbst"),
    ("Hans-Böckler-Solidaritätsfonds", "Nein", "Keine Personal- oder Honorarkosten"),
]
for row in program_rows:
    programs.append(row)
set_widths(programs, [34, 28, 82])
programs.freeze_panes = "A2"

controls = workbook.create_sheet("Prüfgates")
controls.append(["Gate", "Status", "Nachweis"])
style_header(controls[1])
gate_rows = [
    ("WE AID ist echter Zuwendungsempfänger", "offen", "schriftliche Bestätigung"),
    ("Thomas-Vergütung förderfähig", "offen", "Förderer und WE AID"),
    ("Arbeitgeberkosten belastbar", "offen", "Lohnkalkulation WE AID"),
    ("7 Prozent förderfähig", "offen", "Förderer"),
    ("Reichweitentests förderfähig", "offen", "Förderer"),
    ("Kofinanzierung liquide und bestätigt", "offen", "Konto/Spendenzusage"),
    ("Keine Doppelförderung", "offen", "finale Kostenstellenzuordnung"),
    ("Kein privater Liquiditätsbedarf", "offen", "Abschlags-/Zahlungsplan"),
    ("ALG-I-Übergang geklärt", "offen", "schriftliche Auskunft vor Vertrag"),
]
for row in gate_rows:
    controls.append(row)
for cell in controls["B"][1:]:
    cell.fill = input_fill
set_widths(controls, [44, 18, 56])
controls.freeze_panes = "A2"

for sheet in workbook.worksheets:
    sheet.sheet_view.showGridLines = False
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

workbook.save(OUTPUT)

check = load_workbook(OUTPUT, data_only=False)
assert check.sheetnames == ["Annahmen", "Budgetvarianten", "Programmzuordnung", "Prüfgates"]
assert check["Budgetvarianten"]["B13"].value == "=B11-B12"
print(OUTPUT)
